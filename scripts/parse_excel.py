#!/usr/bin/env python3
"""
ERA-ENGINE — Excel Parser v1.0
Baca file "Sales vs Stock" terbaru → output data/live_data.json
Dijalankan otomatis setiap hari oleh launchd / run_update.sh
"""

import sys
import json
import glob
import os
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl belum terinstall. Jalankan: pip3 install openpyxl")
    sys.exit(1)

# ── Konfigurasi ──────────────────────────────────────────────────────────────

REPO_ROOT   = Path(__file__).resolve().parent.parent
DATA_DIR    = REPO_ROOT / "data"
OUTPUT_FILE = DATA_DIR / "live_data.json"

# Folder tempat Excel dari Melati disimpan (sesuaikan jika berbeda)
EXCEL_FOLDER = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/GHESANG HD DRIVE/ERA-ENGINE DASHBOARD CODE/SALES VS STOCK_EMAIL MELATI"

# Jika folder tidak ditemukan, cari di parent dari script ini
if not EXCEL_FOLDER.exists():
    EXCEL_FOLDER = REPO_ROOT / "SALES VS STOCK_EMAIL MELATI"

CURRENT_MONTH  = "May"          # Nama bulan di kolom 'Month' pada sheet SALES VS STOCK
PREV_MONTH     = "April"
PERIOD_LABEL   = "1 - 28 Mei 2026"
WORKING_DAYS   = 28             # Hari kerja MTD bulan ini
DOS_CRITICAL   = 75

# ── Utility ──────────────────────────────────────────────────────────────────

def excel_date_to_month(serial):
    """Konversi Excel date serial ke string 'YYYY-MM'."""
    try:
        return (date(1899, 12, 30) + timedelta(days=int(float(serial)))).strftime("%Y-%m")
    except Exception:
        return ""

def to_num(val):
    """Konversi nilai apapun ke float, aman."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("Rp", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def fmt_compact(n):
    """Format angka ke string compact: 5.063.100 → '5,1Jt' | 730.000.000.000 → '730,0M'."""
    n = abs(n)
    if n >= 1e12:
        return f"{n/1e12:.1f}T".replace(".", ",")
    if n >= 1e9:
        return f"{n/1e9:.1f}M".replace(".", ",")
    if n >= 1e6:
        return f"{n/1e6:.1f}Jt".replace(".", ",")
    if n >= 1e3:
        return f"{n/1e3:.1f}K".replace(".", ",")
    return str(int(n))

def find_latest_excel(folder: Path) -> Path:
    """Cari file .xlsx terbaru di folder."""
    files = list(folder.glob("*.xlsx")) + list(folder.glob("*.XLSX"))
    if not files:
        raise FileNotFoundError(f"Tidak ada file .xlsx di: {folder}")
    return max(files, key=lambda f: f.stat().st_mtime)

# ── Sheet Readers ─────────────────────────────────────────────────────────────

def col_idx(col_letter):
    """Konversi kolom letter (A, B, AA, NB...) ke index 0-based."""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result - 1

def read_sales_vs_stock(ws):
    """
    Sheet: SALES VS STOCK
    Kolom: site_code, site_desc, Month, item_group_desc, brand_name,
           Model Type Detail, Model Category, Qty
    Return: store_qty, brand_qty, category_qty, sku_qty
    """
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1

    site_code_col  = headers.get("site_code", 0)
    site_desc_col  = headers.get("site_desc", 1)
    month_col      = headers.get("Month", 3)
    group_col      = headers.get("item_group_desc", 4)
    brand_col      = headers.get("brand_name", 8)
    qty_col        = headers.get("Qty", 9)
    model_det_col  = headers.get("Model Type Detail", 28)  # col 29
    model_cat_col  = headers.get("Model Category", 29)     # col 30

    store_qty    = {}  # {code: {april: n, may: n, desc: s}}
    brand_qty    = {}  # {brand: {april: n, may: n}}
    category_qty = {}  # {category_group: {april: n, may: n}}
    sku_qty      = {}  # {(brand, model_detail, storage): {april: n, may: n}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[site_code_col]:
            continue

        code  = str(row[site_code_col] or "").strip()
        desc  = str(row[site_desc_col] or "").strip()
        month = str(row[month_col] or "").strip()
        group = str(row[group_col] or "").strip()
        brand = str(row[brand_col] or "").strip()
        qty   = to_num(row[qty_col] if qty_col < len(row) else 0)

        if not code or qty == 0:
            continue

        is_april = month.lower() == PREV_MONTH.lower()
        is_may   = month.lower() == CURRENT_MONTH.lower()

        # Per-store
        if code not in store_qty:
            store_qty[code] = {"april": 0, "may": 0, "desc": desc}
        if is_april: store_qty[code]["april"] += qty
        elif is_may: store_qty[code]["may"]   += qty

        # Per-brand
        brand_key = brand.upper() or "LAINNYA"
        if brand_key not in brand_qty:
            brand_qty[brand_key] = {"april": 0, "may": 0}
        if is_april: brand_qty[brand_key]["april"] += qty
        elif is_may: brand_qty[brand_key]["may"]   += qty

        # Category (gunakan prefix item_group)
        cat = group.split("-")[0].strip().upper() if "-" in group else group.upper()
        if cat not in category_qty:
            category_qty[cat] = {"april": 0, "may": 0}
        if is_april: category_qty[cat]["april"] += qty
        elif is_may: category_qty[cat]["may"]   += qty

        # Per-SKU (brand + model + storage) — hanya baris dengan model valid
        model_det = str(row[model_det_col] if model_det_col < len(row) else "").strip()
        model_cat = str(row[model_cat_col] if model_cat_col < len(row) else "").strip()
        if (model_det and model_det not in ("0", "-", "")
                and model_cat and model_cat not in ("0", "-", "")):
            sku_key = (brand_key, model_det.upper(), model_cat.upper())
            if sku_key not in sku_qty:
                sku_qty[sku_key] = {"april": 0, "may": 0}
            if is_april: sku_qty[sku_key]["april"] += qty
            elif is_may: sku_qty[sku_key]["may"]   += qty

    return store_qty, brand_qty, category_qty, sku_qty


def read_by_store(ws):
    """
    Sheet: BY STORE
    Header multi-row (row 2 = period labels, row 3 = category labels)
    Return: {site_code: {channel, tsh, april_revenue, may_revenue, est_revenue, dos}}
    """
    # (row2 dan row3 dibaca inline di bawah via ws[2] dan ws.cell)

    # Kolom tetap (A-H): baca dari row 2 sebagai dict {text: col_index}
    fixed = {}
    dos_col     = None
    est_col     = None
    target_col  = None

    for cell in ws[2]:
        val = str(cell.value or "").strip()
        col_i = cell.column - 1
        if val in ("Status","BEP","KET SSSG","Channel","TSH","TSH YTD","Code","Store Name"):
            fixed[val] = col_i
        elif val == "DOS":
            dos_col = col_i
        elif val == "Est":
            cat3 = str(ws.cell(row=3, column=cell.column).value or "").strip()
            if cat3 == "":
                est_col = col_i
        elif "Target Mei" in val:
            target_col = col_i

    code_col    = fixed.get("Code", 6)
    name_col    = fixed.get("Store Name", 7)
    channel_col = fixed.get("Channel", 3)
    tsh_col     = fixed.get("TSH", 4)

    # Cari kolom total April dan May
    # openpyxl return datetime untuk sel tanggal
    from datetime import datetime as dt

    april_total_col = None
    may_total_col   = None

    for cell in ws[2]:  # row 2 = period headers
        val = cell.value
        col_i = cell.column - 1
        # Ambil kategori sub-header dari row 3
        cat_cell = ws.cell(row=3, column=cell.column).value
        cat = str(cat_cell or "").strip()
        if isinstance(val, dt):
            if val.year == 2026 and val.month == 4 and cat == "":
                april_total_col = col_i
            elif val.year == 2026 and val.month == 5 and cat == "":
                may_total_col = col_i

    stores = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[code_col]:
            continue
        code = str(row[code_col] or "").strip()
        if not code or code in ("Code", ""):
            continue

        def get(col):
            if col is None or col >= len(row):
                return 0
            return to_num(row[col])

        apr_rev = get(april_total_col)
        may_rev = get(may_total_col)
        est_rev = get(est_col)
        dos     = get(dos_col)
        target  = get(target_col)

        stores[code] = {
            "siteCode":   code,
            "siteDesc":   str(row[name_col] or "").strip() if name_col < len(row) else "",
            "channel":    str(row[channel_col] or "").strip() if channel_col < len(row) else "",
            "tsh":        str(row[tsh_col] or "").strip() if tsh_col < len(row) else "",
            "april_rev":  apr_rev,
            "may_rev":    may_rev,
            "est_rev":    est_rev,
            "dos":        dos,
            "target":     target,
        }

    return stores


def read_sum_r5(ws):
    """
    Sheet: SUM R5
    Struktur aktual (openpyxl):
      - Header di row 5, kolom C (index 2) = 'LOB & TSH'
      - col[3] = May 2025, col[4] = April 2026, col[5] = 'Target Mei 26'
      - col[6]+ = daily May 2026 (datetime objects)
    Return: {tsh: {apr_2026, may_mtd, target_may}}, totals
    """
    from datetime import datetime as dt

    # Cari header row: baris yang punya 'LOB & TSH' di salah satu cell
    header_row_idx = None
    tsh_col = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        for j, cell in enumerate(row or []):
            if str(cell or "").strip() in ("LOB & TSH", "TSH"):
                header_row_idx = i
                tsh_col = j
                break
        if header_row_idx:
            break

    if header_row_idx is None:
        return {}, {}

    headers = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]

    # Cari kolom berdasarkan type/content
    apr26_col    = None
    target_col   = None
    may2025_col  = None
    may_cols     = []

    for i, h in enumerate(headers):
        if h is None:
            continue
        # Target column = string dengan kata "Target" dan "Mei"
        if isinstance(h, str) and "Target" in h and "Mei" in h:
            target_col = i
        # Date columns (openpyxl return datetime)
        elif isinstance(h, dt):
            if h.year == 2025 and h.month == 5:
                may2025_col = i      # May 2025 total → untuk YoY
            elif h.year == 2026 and h.month == 4:
                apr26_col = i        # April 2026 total
            elif h.year == 2026 and h.month == 5:
                may_cols.append(i)   # Daily May 2026

    # Weekly slices dari kolom harian Mei 2026
    this_week_cols = may_cols[-7:]    if len(may_cols) >= 7  else may_cols
    prev_week_cols = may_cols[-14:-7] if len(may_cols) >= 14 else []

    tsh_data = {}
    totals   = {
        "apr_2026": 0.0, "may_mtd": 0.0, "target_may": 0.0,
        "may_2025": 0.0, "may_this_week": 0.0, "may_prev_week": 0.0
    }

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or tsh_col >= len(row) or not row[tsh_col]:
            continue
        name = str(row[tsh_col]).strip()
        # Filter subtotals, channel names, brand names
        name_lower = name.lower()
        if (not name
            or name_lower in ("grand total", "total", "lob & tsh", "")
            or name_lower.endswith("total")
            or any(k in name_lower for k in (
                "channel", "device", "vas ", "lob ",
                "samsung", "xiaomi", "apple", "ibox", "erafone",
                "erablue", "era & more", "honor", "operator store",
                "accessories", "laptop", "handphone", "tablet",
                "audio", "wearable", "computing", "smart home",
            ))):
            continue

        apr26   = to_num(row[apr26_col])  if apr26_col  and apr26_col  < len(row) else 0.0
        target  = to_num(row[target_col]) if target_col and target_col < len(row) else 0.0
        may_mtd = sum(to_num(row[c]) for c in may_cols if c < len(row))
        may2025 = to_num(row[may2025_col]) if may2025_col and may2025_col < len(row) else 0.0

        tsh_data[name] = {
            "tsh":      name,
            "apr_2026": apr26,
            "may_mtd":  may_mtd,
            "target":   target
        }
        totals["apr_2026"]      += apr26
        totals["may_mtd"]       += may_mtd
        totals["target_may"]    += target
        totals["may_2025"]      += may2025
        totals["may_this_week"] += sum(to_num(row[c]) for c in this_week_cols if c < len(row))
        totals["may_prev_week"] += sum(to_num(row[c]) for c in prev_week_cols if c < len(row))

    return tsh_data, totals


def read_pivot(ws):
    """
    Sheet: Pivot (sheet1 — TSH × PO/NonPO × April/May)
    Return: {tsh: {po_apr, po_may, nonpo_apr, nonpo_may}}
    Fallback: ambil TSH + April + May kolom A,B,C
    """
    pivot = {}

    # Cari baris yang punya 'TSH' di kolom A
    data_start = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0] or "").strip() == "TSH":
            data_start = i + 1
            break

    if not data_start:
        return pivot

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not row[0]:
            break
        name = str(row[0]).strip()
        if not name or "Grand Total" in name:
            continue
        pivot[name] = {
            "po_apr":    to_num(row[1]) if len(row) > 1 else 0,
            "po_may":    to_num(row[2]) if len(row) > 2 else 0,
            "nonpo_apr": 0,
            "nonpo_may": 0
        }

    return pivot

# ── Assemblers ────────────────────────────────────────────────────────────────

def build_stores(store_qty, by_store):
    """Gabungkan data qty (SALES VS STOCK) dengan metadata store (BY STORE)."""
    result = []
    all_codes = set(list(store_qty.keys()) + list(by_store.keys()))

    for code in all_codes:
        sq = store_qty.get(code, {})
        bs = by_store.get(code, {})

        apr_qty = sq.get("april", 0)
        may_qty = sq.get("may", 0)
        dos     = bs.get("dos", 0)

        # Estimasi stock: DOS × (may_qty / working_days)
        daily_rate = may_qty / WORKING_DAYS if may_qty > 0 else 0
        stock_est  = round(dos * daily_rate) if dos > 0 else 0

        # Estimasi projected (full month)
        est = round(may_qty / WORKING_DAYS * 31) if WORKING_DAYS > 0 else 0

        desc = bs.get("siteDesc") or sq.get("desc") or code

        result.append({
            "siteCode": code,
            "siteDesc": desc,
            "channel":  bs.get("channel", ""),
            "tsh":      bs.get("tsh", ""),
            "april":    int(apr_qty),
            "may":      int(may_qty),
            "stock":    int(stock_est),
            "est":      round(est, 2),
            "dos":      round(dos, 2)
        })

    # Sort by may qty desc
    result.sort(key=lambda x: x["may"], reverse=True)
    return result


def build_brand_mix(brand_qty):
    sorted_brands = sorted(brand_qty.items(), key=lambda x: x[1]["may"], reverse=True)
    top2    = [{"label": b, "value": int(d["may"])} for b, d in sorted_brands[:2]]
    others  = sum(d["may"] for _, d in sorted_brands[2:])
    if others > 0:
        top2.append({"label": "Lainnya", "value": int(others)})
    return top2


def build_brand_ranking(sku_qty):
    """Per-device-brand ranking dengan apr + may + growth untuk brand cards."""
    totals = {}
    for (brand, model, storage), d in sku_qty.items():
        if brand not in totals:
            totals[brand] = {"april": 0, "may": 0}
        totals[brand]["april"] += d["april"]
        totals[brand]["may"]   += d["may"]

    result = []
    for brand, d in sorted(totals.items(), key=lambda x: x[1]["may"], reverse=True):
        apr, may = d["april"], d["may"]
        if may == 0:
            continue
        growth = round((may - apr) / apr * 100, 1) if apr > 0 else 0.0
        result.append({
            "brand":  brand,
            "may":    int(may),
            "april":  int(apr),
            "growth": growth
        })
    return result


def build_tsh_ranking(tsh_data):
    ranking = [(name, d["may_mtd"]) for name, d in tsh_data.items() if d["may_mtd"] > 0]
    ranking.sort(key=lambda x: x[1], reverse=True)
    return [[name, int(val)] for name, val in ranking[:12]]


def build_trend(pivot_data, tsh_data):
    po_apr = sum(d.get("po_apr", 0) for d in pivot_data.values())
    po_may = sum(d.get("po_may", 0) for d in pivot_data.values())
    # Non-PO = total sales - PO
    total_apr = sum(d["apr_2026"] for d in tsh_data.values())
    total_may = sum(d["may_mtd"]  for d in tsh_data.values())
    nonpo_apr = total_apr - po_apr
    nonpo_may = total_may - po_may
    return [
        {"month": "April", "po": int(po_apr), "nonPo": int(max(0, nonpo_apr))},
        {"month": "May",   "po": int(po_may), "nonPo": int(max(0, nonpo_may))}
    ]


def build_summary(stores, totals):
    may_qty   = sum(s["may"]   for s in stores)
    apr_qty   = sum(s["april"] for s in stores)
    stock     = sum(s["stock"] for s in stores)
    with_dos  = [s["dos"] for s in stores if s["dos"] > 0]
    avg_dos   = round(sum(with_dos) / len(with_dos), 2) if with_dos else 0
    critical  = sum(1 for s in stores if s["dos"] > DOS_CRITICAL)

    return {
        "totalSalesMay":   int(totals.get("may_mtd", 0)),
        "totalSalesApril": int(totals.get("apr_2026", 0)),
        "qtyMay":          int(may_qty),
        "qtyApril":        int(apr_qty),
        "totalStock":      int(stock),
        "avgDos":          avg_dos,
        "criticalStores":  critical,
        "storeCount":      len(stores)
    }


def build_overview(totals, stores):
    target       = int(totals.get("target_may", 0))
    mtd          = int(totals.get("may_mtd", 0))
    apr          = int(totals.get("apr_2026", 0))
    may_2025     = int(totals.get("may_2025", 0))
    this_week    = int(totals.get("may_this_week", 0))
    prev_week    = int(totals.get("may_prev_week", 0))

    # Estimasi akhir bulan
    daily_rate = mtd / WORKING_DAYS if WORKING_DAYS > 0 else 0
    est        = int(daily_rate * 31)

    mtd_pct = round((mtd / target * 100)) if target > 0 else 0
    est_pct = round((est / target * 100)) if target > 0 else 0

    # Timegone (hari berlalu dari total hari kerja bulan)
    tg = round((WORKING_DAYS / 31) * 100)

    # WoW — dari kolom harian Mei (last 7 vs prev 7 working days)
    wow = f"{round((this_week - prev_week) / prev_week * 100):+}%" if prev_week > 0 else "~"

    # MoM — Mei 2026 vs April 2026
    mom = f"{round((mtd - apr) / apr * 100):+}%" if apr > 0 else "~"

    # YoY — Mei 2026 vs Mei 2025 (jika kolom tersedia di sheet)
    yoy = f"{round((mtd - may_2025) / may_2025 * 100):+}%" if may_2025 > 0 else "~"

    # AVG qty per active store
    total_qty     = sum(s["may"] for s in stores)
    active_stores = max(len([s for s in stores if s["may"] > 0]), 1)
    avg_qty       = int(total_qty / active_stores)

    # ASP — Average Selling Price = total revenue / total qty
    asp = int(mtd / total_qty) if total_qty > 0 else 0

    return {
        "title":         "Sales Performance - B2C Region 5",
        "periodLabel":   PERIOD_LABEL,
        "timegone":      tg,
        "underTimegone": max(0, 100 - mtd_pct),
        "targetUnits":   target,
        "mtdUnits":      mtd,
        "estUnits":      est,
        "rings": [
            {"label": "MTD", "value": min(mtd_pct, 100), "tone": "green" if mtd_pct >= 100 else "amber" if mtd_pct >= 80 else "red"},
            {"label": "EST", "value": min(est_pct, 100), "tone": "green" if est_pct >= 100 else "amber" if est_pct >= 80 else "red"}
        ],
        "quickStats": [
            {"label": "WoW",  "value": wow},
            {"label": "MoM",  "value": mom},
            {"label": "YoY",  "value": yoy},
            {"label": "AVG",  "value": f"{avg_qty:,}".replace(",", ".") + " u/toko"},
            {"label": "ASP",  "value": "Rp " + fmt_compact(asp)}
        ],
        "segments": [],       # diisi dari BY STORE aggregate per TSH/channel
        "categories": []      # diisi dari category_qty
    }


def build_stock_summary(stores, summary):
    return [
        ["Total Stock",   f"{summary['totalStock']:,} unit".replace(",", ".")],
        ["DOS Rata-rata",  f"{summary['avgDos']:.1f}".replace(".", ",") + " hari"],
        ["Toko Kritis",    f"{summary['criticalStores']} toko"],
        ["Active Stores",  f"{summary['storeCount']} toko"]
    ]


def build_stock_stack(stores):
    by_tsh = {}
    for s in stores:
        tsh = s["tsh"] or "UNKNOWN"
        if tsh not in by_tsh:
            by_tsh[tsh] = {"tsh": tsh, "erafone": 0, "ibox": 0, "eramore": 0, "erablue": 0, "samsung": 0, "xiaomi": 0}
        ch = (s["channel"] or "").lower()
        stock = s["stock"]
        if "ibox"   in ch:    by_tsh[tsh]["ibox"]    += stock
        elif "more"  in ch:   by_tsh[tsh]["eramore"]  += stock
        elif "blue"  in ch:   by_tsh[tsh]["erablue"]  += stock
        elif "samsung" in ch: by_tsh[tsh]["samsung"]  += stock
        elif "xiaomi"  in ch: by_tsh[tsh]["xiaomi"]   += stock
        else:                 by_tsh[tsh]["erafone"]  += stock
    return sorted(by_tsh.values(), key=lambda x: sum(x[k] for k in ["erafone","ibox","eramore","erablue","samsung","xiaomi"]), reverse=True)[:8]


def build_risk_table(stores):
    risky = [s for s in stores if s["dos"] > DOS_CRITICAL and s["stock"] > 0]
    risky.sort(key=lambda x: x["dos"], reverse=True)
    return [[s["siteDesc"], s["channel"], s["stock"], s["dos"], s["tsh"]] for s in risky[:12]]


DEVICE_BRANDS = {"APPLE", "SAMSUNG", "XIAOMI", "OPPO", "VIVO", "REALME",
                  "INFINIX", "TECNO", "HONOR", "NOTHING", "ITEL", "POCO"}

def _is_storage(s):
    """True jika string tampak seperti kapasitas storage: '128 GB', '8/256 GB', '1 TB'."""
    s = s.upper()
    return "GB" in s or "TB" in s

def build_model_table(sku_qty):
    """Top phone models dari SALES VS STOCK, sorted by qty May.
    Filter: brand = device brand DAN storage mengandung GB/TB."""
    rows = [
        (brand, model, storage, int(d["april"]), int(d["may"]))
        for (brand, model, storage), d in sku_qty.items()
        if d["may"] > 0
        and brand in DEVICE_BRANDS
        and _is_storage(storage)
    ]
    return sorted(rows, key=lambda x: x[4], reverse=True)[:50]


def build_accessories(category_qty):
    cat_map = {
        "MOBILE": "Shopping Bag",
        "ACC":    "Accessories",
        "AUDIO":  "Audio Products",
        "WEAR":   "Wearable",
        "CASE":   "Case & Sleeve",
        "CHRGR":  "Charger",
        "LAPTOP": "Laptop Accessories",
    }
    result = []
    for cat, qty in sorted(category_qty.items(), key=lambda x: x[1]["may"], reverse=True):
        label = next((v for k, v in cat_map.items() if k in cat.upper()), cat)
        if qty["may"] > 0:
            result.append({"label": label, "value": int(qty["may"])})
    return result[:8] if result else [{"label": "Data belum tersedia", "value": 0}]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"[ERA-ENGINE] Mulai parsing — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Cari Excel terbaru
    try:
        excel_path = find_latest_excel(EXCEL_FOLDER)
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    print(f"[ERA-ENGINE] File: {excel_path.name}")

    # Load workbook (read_only untuk performa)
    print("[ERA-ENGINE] Loading workbook...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    sheet_names = wb.sheetnames
    print(f"[ERA-ENGINE] Sheets ditemukan: {sheet_names}")

    # Cari sheet dengan nama yang sesuai
    def get_sheet(candidates):
        for name in candidates:
            for sn in sheet_names:
                if name.lower() in sn.lower():
                    return wb[sn]
        return None

    ws_svs     = get_sheet(["SALES VS STOCK"])
    ws_store   = get_sheet(["BY STORE", "STORE"])
    ws_sum     = get_sheet(["SUM R5", "SUM", "SUMMARY"])
    ws_pivot1  = get_sheet(["Pivot "])   # Pivot dengan spasi = Pivot sheet 1

    if not ws_svs:
        print("WARNING: Sheet 'SALES VS STOCK' tidak ditemukan")
    if not ws_store:
        print("WARNING: Sheet 'BY STORE' tidak ditemukan")
    if not ws_sum:
        print("WARNING: Sheet 'SUM R5' tidak ditemukan")

    # Parse setiap sheet
    print("[ERA-ENGINE] Parsing SALES VS STOCK...")
    store_qty, brand_qty, category_qty, sku_qty = read_sales_vs_stock(ws_svs) if ws_svs else ({}, {}, {}, {})

    print("[ERA-ENGINE] Parsing BY STORE...")
    by_store = read_by_store(ws_store) if ws_store else {}

    print("[ERA-ENGINE] Parsing SUM R5...")
    tsh_data, totals = read_sum_r5(ws_sum) if ws_sum else ({}, {})

    print("[ERA-ENGINE] Parsing Pivot...")
    pivot_data = read_pivot(ws_pivot1) if ws_pivot1 else {}

    wb.close()

    # Assemble output
    print("[ERA-ENGINE] Assembling dashboard data...")
    stores         = build_stores(store_qty, by_store)
    brand_mix      = build_brand_mix(brand_qty)
    brand_ranking  = build_brand_ranking(sku_qty)
    tsh_ranking  = build_tsh_ranking(tsh_data)
    trend        = build_trend(pivot_data, tsh_data)
    summary      = build_summary(stores, totals)
    overview     = build_overview(totals, stores)
    stock_summ   = build_stock_summary(stores, summary)
    stock_stack  = build_stock_stack(stores)
    risk         = build_risk_table(stores)
    accessories  = build_accessories(category_qty)

    # Highlights
    sorted_by_may = sorted(stores, key=lambda x: x["may"], reverse=True)
    sorted_by_dos = sorted([s for s in stores if s["stock"] > 50], key=lambda x: x["dos"], reverse=True)
    best = sorted_by_may[0] if sorted_by_may else {}
    dead = sorted_by_dos[0] if sorted_by_dos else {}
    highlights = {
        "bestSeller": {
            "title": "Best Seller Store",
            "model": best.get("siteDesc", "-"),
            "value": f"{best.get('may', 0):,} unit".replace(",", "."),
            "note":  f"Toko dengan qty {CURRENT_MONTH} tertinggi."
        },
        "deadStock": {
            "title": "DOS Kritis",
            "model": dead.get("siteDesc", "-"),
            "value": f"DOS {dead.get('dos', 0):.1f} hari",
            "note":  "Butuh intervensi sell-out segera."
        }
    }

    model_table = [
        [brand, model, storage, apr, may, 0, 0]
        for brand, model, storage, apr, may in build_model_table(sku_qty)
    ]

    payload = {
        "generated_at": datetime.now().isoformat(),
        "source_file":  excel_path.name,
        "summary":      summary,
        "overview":     overview,
        "trend":        trend,
        "tshRanking":   tsh_ranking,
        "stores":       stores,
        "brandMix":     brand_mix,
        "brandRanking": brand_ranking,
        "modelTable":   model_table,
        "highlights":   highlights,
        "stockSummary": stock_summ,
        "stockStack":   stock_stack,
        "risk":         risk,
        "accessories":  accessories
    }

    # Tulis output
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    size_kb = OUTPUT_FILE.stat().st_size // 1024
    print(f"[ERA-ENGINE] ✅ Output: {OUTPUT_FILE} ({size_kb} KB)")
    print(f"[ERA-ENGINE] Stores: {len(stores)} | Brands: {len(brand_mix)} | TSH: {len(tsh_ranking)}")


if __name__ == "__main__":
    main()
