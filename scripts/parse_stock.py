"""
parse_stock.py — ERA-ENGINE Stock Intelligence Parser
Reads DATA STOCK.xlsx (Raw Data sheet) and returns stockDetail dict
for embedding in live_data.json.

Column layout (0-indexed):
  [0]  Code Plant
  [1]  Desc
  [3]  Month       → "Stock" | "Stock DC" | "Jan".."May"
  [7]  Type 2      → model name
  [8]  brand_name
  [9]  Qty
  [15] Bu Category (1MULTIBRAND, 2IBOX, etc.)
  [19] Channel     (ERAFONE, IBOX, etc.)
  [20] HOT         (area supervisor name)
  [21] TSH         (area manager name)
  [26] Category Device (model family: A6X, NOTE 15, etc.)
"""

import os
from collections import defaultdict
from datetime import datetime

WORKING_DAYS = 28   # May 2026 working days — update each month

STOCK_FILE = os.path.join(
    os.path.dirname(__file__),
    "..",
    "STOCK",
    "DATA STOCK.xlsx",
)

DEVICE_BRANDS = {
    "APPLE", "SAMSUNG", "XIAOMI", "OPPO", "VIVO", "Vivo",
    "REALME", "Realme", "INFINIX", "Infinix", "TECNO",
    "HONOR", "Honor", "NOTHING", "ITEL", "POCO",
    "IQOO", "MOTOROLA", "HUAWEI", "SHARP", "ADVAN",
}


def dos_status(dos):
    if dos > 90:
        return "danger"
    if dos > 60:
        return "warn"
    return "safe"


def clean(v):
    return str(v).strip() if v else ""


def build_stock_detail():
    """Parse DATA STOCK.xlsx → return stockDetail dict, or {} on failure."""
    try:
        import openpyxl
    except ImportError:
        print("[parse_stock] openpyxl not available — skipping stock detail")
        return {}

    stock_path = STOCK_FILE
    if not os.path.exists(stock_path):
        print(f"[parse_stock] {stock_path} not found — skipping")
        return {}

    print(f"[parse_stock] Reading {os.path.basename(stock_path)} …")
    try:
        wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[parse_stock] Cannot open workbook: {e}")
        return {}

    if "Raw Data" not in wb.sheetnames:
        print("[parse_stock] 'Raw Data' sheet not found")
        wb.close()
        return {}

    ws = wb["Raw Data"]

    # Accumulators
    brand_stock = defaultdict(int)
    brand_may   = defaultdict(int)
    brand_apr   = defaultdict(int)

    tsh_stock   = defaultdict(int)
    tsh_may     = defaultdict(int)

    bu_cat      = defaultdict(int)
    channel_stk = defaultdict(int)
    cat_dev     = defaultdict(int)

    model_data  = defaultdict(lambda: {"brand": "", "stock": 0, "may": 0})

    total_stock = 0
    total_may   = 0

    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        month  = clean(row[3])
        qty    = row[9]
        brand  = clean(row[8])
        type2  = clean(row[7])
        tsh    = clean(row[21])
        bucat  = clean(row[15])
        chan   = clean(row[19])
        catdev = clean(row[26])

        if not qty or qty == 0:
            continue

        if month == "Stock":
            brand_stock[brand]  += qty
            tsh_stock[tsh]      += qty
            bu_cat[bucat]       += qty
            channel_stk[chan]   += qty
            cat_dev[catdev]     += qty
            model_data[type2]["brand"]  = brand
            model_data[type2]["stock"] += qty
            total_stock += qty

        elif month == "May":
            brand_may[brand] += qty
            tsh_may[tsh]     += qty
            model_data[type2]["may"] += qty
            total_may += qty

        elif month == "Apr":
            brand_apr[brand] += qty

    wb.close()
    print(f"[parse_stock] Total stock={total_stock:,}  May sales={total_may:,}")

    # ── Brand stock list ──────────────────────────────────────────────
    brand_list = []
    for b, stk in sorted(brand_stock.items(), key=lambda x: -x[1]):
        if b not in DEVICE_BRANDS:
            continue
        may  = brand_may.get(b, 0)
        apr  = brand_apr.get(b, 0)
        daily = may / WORKING_DAYS if may > 0 else 0
        dos  = round(stk / daily, 1) if daily > 0 else 9999
        brand_list.append({
            "brand":    b,
            "stock":    stk,
            "maySales": may,
            "aprSales": apr,
            "dos":      dos,
            "status":   dos_status(dos),
        })

    # ── TSH stock list ────────────────────────────────────────────────
    tsh_list = []
    for t, stk in sorted(tsh_stock.items(), key=lambda x: -x[1]):
        if not t or t in ("", "None"):
            continue
        may   = tsh_may.get(t, 0)
        daily = may / WORKING_DAYS if may > 0 else 0
        dos   = round(stk / daily, 1) if daily > 0 else 9999
        tsh_list.append({
            "tsh":    t,
            "stock":  stk,
            "may":    may,
            "dos":    dos,
            "status": dos_status(dos),
        })

    # ── Bu Category list ──────────────────────────────────────────────
    bucat_list = [
        {"label": k, "stock": v}
        for k, v in sorted(bu_cat.items(), key=lambda x: -x[1])
        if k
    ]

    # ── Channel list ──────────────────────────────────────────────────
    channel_list = [
        {"channel": k, "stock": v}
        for k, v in sorted(channel_stk.items(), key=lambda x: -x[1])
        if k
    ]

    # ── Category Device list ──────────────────────────────────────────
    catdev_list = [
        {"label": k, "stock": v}
        for k, v in sorted(cat_dev.items(), key=lambda x: -x[1])[:30]
        if k
    ]

    # ── Model detail list (top 200 by stock) ─────────────────────────
    model_list = []
    for model, d in sorted(model_data.items(), key=lambda x: -x[1]["stock"])[:200]:
        if not model:
            continue
        stk  = d["stock"]
        may  = d["may"]
        daily = may / WORKING_DAYS if may > 0 else 0
        dos  = round(stk / daily, 1) if daily > 0 else (9999 if stk > 0 else 0)
        model_list.append({
            "model":  model,
            "brand":  d["brand"],
            "stock":  stk,
            "may":    may,
            "dos":    dos,
            "status": dos_status(dos) if dos < 9999 else "danger",
        })

    # ── Summary counts ────────────────────────────────────────────────
    kritis_brands  = sum(1 for x in brand_list if x["status"] == "danger")
    waspada_brands = sum(1 for x in brand_list if x["status"] == "warn")
    aman_brands    = sum(1 for x in brand_list if x["status"] == "safe")
    kritis_models  = sum(1 for x in model_list if x["status"] == "danger")
    total_may_device = sum(b["maySales"] for b in brand_list)
    avg_dos_device   = round(sum(b["dos"] for b in brand_list if b["dos"] < 9999) /
                             max(len([b for b in brand_list if b["dos"] < 9999]), 1), 1)

    return {
        "totalStock":    total_stock,
        "totalMaySales": total_may,
        "avgDos":        avg_dos_device,
        "kritisCount":   kritis_brands,
        "waspadaCount":  waspada_brands,
        "amanCount":     aman_brands,
        "kritisModels":  kritis_models,
        "brandStock":    brand_list,
        "tshStock":      tsh_list,
        "buCatStock":    bucat_list,
        "channelStock":  channel_list,
        "catDevStock":   catdev_list,
        "modelDetail":   model_list,
        "sourceFile":    os.path.basename(stock_path),
        "parsedAt":      datetime.now().isoformat(),
    }


if __name__ == "__main__":
    import json
    result = build_stock_detail()
    out_path = os.path.join(os.path.dirname(__file__), "..", "data", "stock_detail.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[parse_stock] Written → {out_path}")
